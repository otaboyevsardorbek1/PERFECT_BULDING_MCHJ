import pandas as pd
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime, date
import os
from typing import List, Dict, Any
from config import EXCEL_REPORTS_DIR
import io

def create_excel_report(data: List[Dict], report_type: str, title: str = None) -> str:
    """Excel hisobot yaratish"""
    
    # Filename yaratish
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"{report_type}_{timestamp}.xlsx"
    filepath = os.path.join(EXCEL_REPORTS_DIR, filename)
    
    # Workbook yaratish
    wb = Workbook()
    ws = wb.active
    ws.title = report_type.capitalize()
    
    # Sarlavha yaratish
    if title:
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = title
        title_cell.font = Font(size=16, bold=True)
        title_cell.alignment = Alignment(horizontal='center')
    
    # Ma'lumotlarni yozish
    if data:
        # Ustun sarlavhalari
        headers = list(data[0].keys())
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
        # Ma'lumotlar qatorlari
        for row_idx, row_data in enumerate(data, 4):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                ws.cell(row=row_idx, column=col_idx, value=value)
        
        # Stil berish
        for col_idx in range(1, len(headers) + 1):
            col_letter = get_column_letter(col_idx)
            ws.column_dimensions[col_letter].auto_size = True
        
        # Chegaralar
        thin_border = Border(left=Side(style='thin'), 
                           right=Side(style='thin'), 
                           top=Side(style='thin'), 
                           bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=3, max_row=len(data)+3, min_col=1, max_col=len(headers)):
            for cell in row:
                cell.border = thin_border
    
    # Sana va vaqt
    ws.cell(row=len(data)+5, column=1, value="Yaratilgan sana:")
    ws.cell(row=len(data)+5, column=2, value=datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    
    # Faylni saqlash
    wb.save(filepath)
    return filepath

def create_warehouse_excel_report(raw_materials: List[Dict], products: List[Dict]) -> str:
    """Ombor holati bo'yicha Excel hisobot"""
    
    wb = Workbook()
    
    # 1. Xom ashyolar
    ws1 = wb.active
    ws1.title = "Xom ashyolar"
    
    # Sarlavha
    ws1.merge_cells('A1:F1')
    ws1['A1'] = "XOM ASHYOLAR HOLATI"
    ws1['A1'].font = Font(size=14, bold=True)
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    # Ustunlar
    headers = ["№", "Nomi", "Birlik", "Jami", "Minimal", "Narx", "Holat"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Ma'lumotlar
    for idx, material in enumerate(raw_materials, 1):
        row = idx + 3
        status = "✅ Yetarli" if material['current_stock'] > material['min_stock'] else "⚠️ Yetarli emas"
        
        ws1.cell(row=row, column=1, value=idx)
        ws1.cell(row=row, column=2, value=material['name'])
        ws1.cell(row=row, column=3, value=material['unit'])
        ws1.cell(row=row, column=4, value=material['current_stock'])
        ws1.cell(row=row, column=5, value=material['min_stock'])
        ws1.cell(row=row, column=6, value=material['price_per_unit'])
        ws1.cell(row=row, column=7, value=status)
    
    # 2. Mahsulotlar
    ws2 = wb.create_sheet("Mahsulotlar")
    
    # Sarlavha
    ws2.merge_cells('A1:E1')
    ws2['A1'] = "TAYYOR MAHSULOTLAR"
    ws2['A1'].font = Font(size=14, bold=True)
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    # Ustunlar
    headers = ["№", "Nomi", "Birlik", "Sotish narxi", "Xarajat", "Foyda %"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws2.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    # Ma'lumotlar
    for idx, product in enumerate(products, 1):
        row = idx + 3
        profit_margin = ((product['selling_price'] - product['production_cost']) / product['production_cost'] * 100) if product['production_cost'] > 0 else 0
        
        ws2.cell(row=row, column=1, value=idx)
        ws2.cell(row=row, column=2, value=product['name'])
        ws2.cell(row=row, column=3, value=product['unit'])
        ws2.cell(row=row, column=4, value=product['selling_price'])
        ws2.cell(row=row, column=5, value=product['production_cost'])
        ws2.cell(row=row, column=6, value=f"{profit_margin:.1f}%")
    
    # Faylni saqlash
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"warehouse_report_{timestamp}.xlsx"
    filepath = os.path.join(EXCEL_REPORTS_DIR, filename)
    wb.save(filepath)
    
    return filepath

def create_financial_excel_report(data: Dict, period: str) -> str:
    """Moliya hisoboti uchun Excel"""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Moliya hisoboti"
    
    # Asosiy sarlavha
    ws.merge_cells('A1:D1')
    ws['A1'] = f"MOLIYA HISOBOTI - {period.upper()}"
    ws['A1'].font = Font(size=16, bold=True, color="FFFFFF")
    ws['A1'].alignment = Alignment(horizontal='center')
    ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Daromadlar bo'limi
    ws['A3'] = "DAROMADLAR"
    ws['A3'].font = Font(size=12, bold=True)
    
    income_items = [
        ("Sotuvdan daromad", data.get('total_sales_amount', 0)),
        ("Boshqa daromadlar", data.get('other_income', 0)),
    ]
    
    for idx, (item, amount) in enumerate(income_items, 4):
        ws.cell(row=idx, column=1, value=item)
        ws.cell(row=idx, column=2, value=amount)
    
    total_income = sum([item[1] for item in income_items])
    ws.cell(row=len(income_items)+4, column=1, value="JAMI DAROMAD")
    ws.cell(row=len(income_items)+4, column=2, value=total_income)
    ws.cell(row=len(income_items)+4, column=1).font = Font(bold=True)
    ws.cell(row=len(income_items)+4, column=2).font = Font(bold=True)
    
    # Xarajatlar bo'limi
    start_row = len(income_items) + 7
    ws.cell(row=start_row, column=1, value="XARAJATLAR")
    ws.cell(row=start_row, column=1).font = Font(size=12, bold=True)
    
    expense_items = [
        ("Ishlab chiqarish xarajatlari", data.get('production_costs', 0)),
        ("Maosh to'lovlari", data.get('salary_costs', 0)),
        ("Kommunal to'lovlar", data.get('utility_costs', 0)),
        ("Ijaralar", data.get('rent_costs', 0)),
        ("Boshqa xarajatlar", data.get('other_expenses', 0)),
    ]
    
    for idx, (item, amount) in enumerate(expense_items, start_row+1):
        ws.cell(row=idx, column=1, value=item)
        ws.cell(row=idx, column=2, value=amount)
    
    total_expenses = sum([item[1] for item in expense_items])
    total_row = start_row + len(expense_items) + 1
    ws.cell(row=total_row, column=1, value="JAMI XARAJAT")
    ws.cell(row=total_row, column=2, value=total_expenses)
    ws.cell(row=total_row, column=1).font = Font(bold=True)
    ws.cell(row=total_row, column=2).font = Font(bold=True)
    
    # Sof foyda
    profit_row = total_row + 2
    net_profit = total_income - total_expenses
    profit_margin = (net_profit / total_income * 100) if total_income > 0 else 0
    
    ws.cell(row=profit_row, column=1, value="SOF FOYDA")
    ws.cell(row=profit_row, column=2, value=net_profit)
    ws.cell(row=profit_row, column=1).font = Font(bold=True, size=14)
    ws.cell(row=profit_row, column=2).font = Font(bold=True, size=14)
    
    ws.cell(row=profit_row+1, column=1, value="Foyda marjasi")
    ws.cell(row=profit_row+1, column=2, value=f"{profit_margin:.2f}%")
    
    # Formatlash
    for col in ['A', 'B']:
        ws.column_dimensions[col].width = 30
    
    for row in range(1, profit_row+2):
        for col in [1, 2]:
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal='left')
            if col == 2 and row > 3:
                cell.number_format = '#,##0'
    
    # Faylni saqlash
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"financial_report_{period}_{timestamp}.xlsx"
    filepath = os.path.join(EXCEL_REPORTS_DIR, filename)
    wb.save(filepath)
    
    return filepath

def create_employee_report(employees: List[Dict], work_hours_data: Dict, salary_data: Dict) -> str:
    """Xodimlar hisoboti"""
    
    wb = Workbook()
    
    # 1. Xodimlar ro'yxati
    ws1 = wb.active
    ws1.title = "Xodimlar"
    
    ws1.merge_cells('A1:G1')
    ws1['A1'] = "XODIMLAR RO'YXATI"
    ws1['A1'].font = Font(size=14, bold=True)
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    headers = ["ID", "F.I.Sh", "Lavozim", "Bo'lim", "Telefon", "Ishga kirgan", "Maosh"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws1.cell(row=3, column=col_idx, value=header)
        cell.font = Font(bold=True)
    
    for idx, emp in enumerate(employees, 1):
        row = idx + 3
        ws1.cell(row=row, column=1, value=emp['id'])
        ws1.cell(row=row, column=2, value=emp['full_name'])
        ws1.cell(row=row, column=3, value=emp['position'])
        ws1.cell(row=row, column=4, value=emp['department'])
        ws1.cell(row=row, column=5, value=emp['phone_number'])
        ws1.cell(row=row, column=6, value=emp['hire_date'].strftime("%Y-%m-%d") if isinstance(emp['hire_date'], datetime) else emp['hire_date'])
        ws1.cell(row=row, column=7, value=emp['salary'])
    
    # 2. Ish vaqtlari
    ws2 = wb.create_sheet("Ish vaqtlari")
    
    ws2.merge_cells('A1:E1')
    ws2['A1'] = "ISH VAQTLARI HISOBOTI"
    ws2['A1'].font = Font(size=14, bold=True)
    ws2['A1'].alignment = Alignment(horizontal='center')
    
    # Ish vaqtlari ma'lumotlarini yozish
    # ...
    
    # 3. Maosh to'lovlari
    ws3 = wb.create_sheet("Maosh to'lovlari")
    
    ws3.merge_cells('A1:H1')
    ws3['A1'] = "MAOSH TO'LOVLARI"
    ws3['A1'].font = Font(size=14, bold=True)
    ws3['A1'].alignment = Alignment(horizontal='center')
    
    # Maosh to'lovlari ma'lumotlarini yozish
    # ...
    
    # Faylni saqlash
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"employee_report_{timestamp}.xlsx"
    filepath = os.path.join(EXCEL_REPORTS_DIR, filename)
    wb.save(filepath)
    
    return filepath